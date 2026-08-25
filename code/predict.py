#!/usr/bin/env python
"""One-click final prediction: v3 benchmark xlsx template -> final_predictions.xlsx

Pipeline (each step is the repo script of the same name, run as a subprocess
with the current interpreter):
  1. data_pipeline/parse_v3_xlsx.py  -- v3 xlsx -> data/benchmark_mAbs_v3.csv
  2. training/infer_benchmark.py     -- 3-seed checkpoints, Borda ensemble
  3. this script                     -- fills the prediction-rank column of a
                                        template copy via openpyxl (format-preserving)

Expected runtime on CPU: ~5-10 min (ESM-2 antigen extraction dominates).

Run from anywhere:
  python code/predict.py [--template <v3 benchmark xlsx>] [--out final_predictions.xlsx]
"""
import argparse
import subprocess
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

CODE_DIR = Path(__file__).resolve().parent
DEFAULT_TEMPLATE = CODE_DIR / "data" / "benchmark_v3.xlsx"  # v3 模板；不入库，需手动放置
DEFAULT_OUT = CODE_DIR.parent / "final_predictions.xlsx"
RANKING_CSV = CODE_DIR / "predictions" / "mAbs_ranking.csv"
# header labels in the v3 template (sic: the sheet spells "Sequene")
COL_GROUP, COL_SEQID, COL_PRED = "Group", "Sequene ID", "Team X"


def run_step(script, *script_args):
    cmd = [sys.executable, str(CODE_DIR / script), *map(str, script_args)]
    print(f"+ {' '.join(cmd)}", flush=True)
    subprocess.run(cmd, cwd=CODE_DIR, check=True)


def fill_team_x(template, ranking_csv, out_path):
    ranks = pd.read_csv(ranking_csv)
    pred = {(int(r.group), str(r.seq_id)): int(r.pred_rank)
            for r in ranks.itertuples()}
    wb = load_workbook(template)
    sheets = [n for n in wb.sheetnames if "mab" in n.lower()] or wb.sheetnames[:1]
    ws = wb[sheets[0]]
    header = {str(c.value).strip(): c.column for c in ws[1] if c.value is not None}
    for name in (COL_GROUP, COL_SEQID, COL_PRED):
        if name not in header:
            raise SystemExit(f"template header missing '{name}': {header}")
    filled, group_val = 0, None
    for row in range(2, ws.max_row + 1):
        g = ws.cell(row=row, column=header[COL_GROUP]).value
        if g is not None:  # merged cells: only the first row of a block has a value
            group_val = int(g)
        seq_id = ws.cell(row=row, column=header[COL_SEQID]).value
        if seq_id is None:
            continue
        key = (group_val, str(seq_id).strip())
        if key not in pred:
            raise SystemExit(f"no prediction for template row {row}: {key}")
        ws.cell(row=row, column=header[COL_PRED]).value = pred[key]
        filled += 1
    if filled != len(pred):
        raise SystemExit(f"filled {filled} template rows but have {len(pred)} predictions")
    for gid, sub in ranks.groupby("group"):
        assert sorted(sub.pred_rank) == list(range(1, len(sub) + 1))
    wb.save(out_path)
    print(f"filled prediction rank for {filled} rows -> {out_path}")


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE)
    ap.add_argument("--out", type=Path, default=DEFAULT_OUT)
    ap.add_argument("--skip-infer", action="store_true",
                    help="reuse the existing ranking CSV (debug)")
    args = ap.parse_args()

    run_step("data_pipeline/parse_v3_xlsx.py", "--xlsx", args.template)
    if not args.skip_infer:
        run_step("training/infer_benchmark.py", "--out", RANKING_CSV)
    fill_team_x(args.template, RANKING_CSV, args.out)
    print("predict.py: done")


if __name__ == "__main__":
    main()
