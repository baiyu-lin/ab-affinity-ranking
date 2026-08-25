#!/usr/bin/env python
"""Generate docs/assets/附表A_消融实验汇总.xlsx (4 sheets).

All numbers are transcribed from the decision records in planning/ and
archive_2026-08-08/docs/ablation-final-report.md — no new experiments.

Run: code/.venv/bin/python docs/make_appendix_xlsx.py
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent / "assets" / "附表A_消融实验汇总.xlsx"

HDR_FILL = PatternFill("solid", fgColor="1E40AF")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
KEY_FILL = PatternFill("solid", fgColor="DCFCE7")   # adopted rows
NEG_FILL = PatternFill("solid", fgColor="FEE2E2")   # rejected/negative rows
BODY_FONT = Font(size=10)
WRAP = Alignment(vertical="center", wrap_text=True)


def sheet(wb, title, headers, rows, note=None, widths=None):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for c in ws[1]:
        c.fill, c.font, c.alignment = HDR_FILL, HDR_FONT, WRAP
    for row in rows:
        ws.append(row)
        r = ws.max_row
        mark = row[-1] if isinstance(row[-1], str) else ""
        for c in ws[r]:
            c.font, c.alignment = BODY_FONT, WRAP
            if "采纳" in mark or "入选" in mark or "最终" in mark:
                c.fill = KEY_FILL
            elif "否决" in mark or "淘汰" in mark or "证伪" in mark:
                c.fill = NEG_FILL
    if note:
        ws.append([])
        ws.append([note])
        ws.cell(ws.max_row, 1).font = Font(size=9, italic=True, color="64748B")
    for i, w in enumerate(widths or [16] * len(headers), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    return ws


wb = Workbook()
wb.remove(wb.active)

# ---- sheet 1: 架构消融 A0-A4 + 稳定化研究 -----------------------------------
sheet(
    wb, "1_架构消融A0-A4",
    ["消融", "臂 / 配方", "AbRank 5 折均值", "抗原留出 fold0", "判定", "依据说明"],
    [
        ["A0 完整性", "线性探针（pooled 512+1280→Linear）", "0.214", "0.068", "否决", "多样抗原折崩塌，无跨抗原迁移"],
        ["A0 完整性", "零样本 IgLM / ProGen2", "≈0 / 仅 kothiwal-DCC 0.322", "—", "否决", "与 FLAb 基准报告一致"],
        ["A0 完整性", "完整模型（交叉注意力+双线性门控）", "0.276", "0.083", "采纳", "较探针 +0.062，5 折胜 4 折"],
        ["A1 抗体塔", "AntiBERTy（冻结）", "0.276 ± 0.064", "0.083", "入选", "更优且更稳定"],
        ["A1 抗体塔", "IgLM（GPU 重提取）", "0.261 ± 0.059", "0.058", "淘汰", "fold0 种子散布 0.094–0.257，三次 epoch-0 早停"],
        ["A2 交互", "交叉注意力（Q=抗原）", "0.276 ± 0.064", "0.083 ± 0.051", "入选", "最严格留出轴上较盲抗原接近翻倍"],
        ["A2 交互", "盲抗原对照", "0.256 ± 0.075", "0.054 ± 0.026", "否决", "抗原条件化收益集中在泛化轴"],
        ["A2 交互", "仅拼接（无交互）", "0.242 ± 0.089", "0.075 ± 0.015", "否决", "AbRank 0/1/3 折落败"],
        ["A3 融合", "双线性门控 vs 纯拼接", "门控 ≥ 拼接", "门控 ≥ 拼接", "采纳", "两评估轴均值均占优"],
        ["A4 数据", "rank_only 排除（factor=0）", "−0.011（vs 0.276）", "—", "否决", "AbRank 合格组大多为 rank-only"],
        ["A4 数据", "rank_only 0.5× 降权", "0.276", "—", "采纳", "SPR-Kd 迁移：DCC 折 0.039→0.350"],
        ["A4b 辅助损失", "λ_aux（VHH margin）0.2 vs 0", "0.263 vs 0.261", "—", "λ_aux=0", "中性，目标保持单一 ListMLE"],
        ["稳定化", "lr 1e-4（融合 3e-4）", "0.250 ± 0.088", "—", "否决", "欠训"],
        ["稳定化", "500 步/轮（原基线）", "0.276 ± 0.064", "—", "否决", "—"],
        ["稳定化", "1500 步/轮", "0.297 ± 0.066", "—", "采纳", "4/5 折获胜，达预设 ±0.02 门槛"],
    ],
    note="协议：3 种子 mean-of-best，10 轮，AdamW lr 3e-4；出处：archive_2026-08-08/docs/ablation-final-report.md。",
    widths=[12, 34, 20, 16, 10, 40],
)

# ---- sheet 2: 输入侧消融 -----------------------------------------------------
sheet(
    wb, "2_输入侧消融",
    ["实验", "指标", "消融臂", "基线（同协议）", "判定", "说明"],
    [
        ["链感知接口 chain_ids", "监控集 Spearman s0/s1/s2", "0.787 / 0.749 / 0.666", "0.804 / 0.766 / 0.700", "否决", "3/3 种子全劣"],
        ["链感知接口 chain_ids", "v3 自评 Group 1（集成）", "0.278", "0.585", "否决", "首轮旧实例检查点"],
        ["链感知接口 chain_ids", "v3 自评 Group 2（集成）", "0.786", "0.842", "否决", "同上"],
        ["链感知接口 chain_ids", "v3 自评 均值", "0.532", "0.714", "否决", "可学习链型偏置破坏冻结表征几何"],
        ["链掩码探针", "完整输入 G1 / G2", "0.593 / 0.845", "—", "—", "最终模型掩码探针（probe_chains.py）"],
        ["链掩码探针", "仅 VH 输入 G1 / G2", "0.382 / 0.833", "—", "—", "VH 承载主导排序信号"],
        ["链掩码探针", "仅 VL 输入 G1 / G2", "0.075 / ≈0（常数）", "—", "—", "Group 2 VL 无组内变异，符合设计"],
        ["混合编码器塔", "—", "—", "—", "不启动", "探针未指向 AntiBERTy 链表征不足"],
    ],
    note="出处：设计文档 §2.6；链感知臂与基线为首轮旧实例检查点，结论不受后续复训影响。",
    widths=[22, 28, 24, 22, 10, 38],
)

# ---- sheet 3: 结构第三塔 ------------------------------------------------------
sheet(
    wb, "3_结构第三塔",
    ["臂", "结构编码路径", "粒度 / 注入点", "s0", "s1", "s2", "均值", "Δ vs 基线", "判定"],
    [
        ["基线（同协议快照）", "—", "—", 0.8042, 0.7663, 0.7003, 0.7569, "—", "—"],
        ["ProstT5 残差", "ProstT5（免结构预测 AA→3Di）", "池化向量残差→B_pool/A_pool", 0.7983, 0.7595, 0.7005, 0.7528, "−0.004", "证伪"],
        ["SaProt 残差", "IgFold→foldseek 3Di→SaProt-650M", "池化向量残差→B_pool", 0.8042, 0.7661, 0.7025, 0.7576, "+0.001", "证伪（零增益）"],
        ["SaProt 交叉注意力", "同上", "池化→8 struct tokens，B_pool 作 Q", 0.7947, 0.7387, 0.6972, 0.7435, "−0.013", "证伪"],
        ["SaProt 逐残基门控 gated_pre", "同上，逐残基 [L,1280]", "抗体塔 FFN 前门控残差，ReZero 零初始化", 0.8012, 0.7710, 0.6957, 0.7560, "−0.001", "证伪（零增益）"],
        ["OOD 补充：ProstT5 臂", "AbRank fold0 抗原留出", "—", 0.2333, 0.2865, 0.2984, 0.2727, "+0.021", "未达 +0.03 门槛"],
        ["OOD 基线", "AbRank fold0 抗原留出", "—", 0.2748, 0.2291, 0.2499, 0.2513, "—", "—"],
        ["Arm0 真实结构预检", "SAbDab 精确匹配", "抗原 23/25 命中；抗体 314/144,140（0.22%）", "—", "—", "—", "—", "—", "真实结构路线不可行"],
    ],
    note="协议：final_all cv=all，监控集 100 组，3 种子 best val_spearman，10 轮 × 1500 步。出处：planning/ 2026-08-09/08-11/08-12 三篇结构臂记录。",
    widths=[24, 30, 34, 9, 9, 9, 9, 11, 16],
)

# ---- sheet 4: v3 自评对照 -----------------------------------------------------
sheet(
    wb, "4_v3自评对照",
    ["模型", "Group 1（SARS-CoV-2）", "Group 2（HER2）", "均值", "备注"],
    [
        ["去泄漏重训（最终模型，3 种子 Borda 集成）", 0.395, 0.938, 0.667, "最终模型"],
        ["　逐种子 G1", "0.287 / 0.304 / 0.155", "—", "—", "种子间离散大"],
        ["　逐种子 G2", "—", "0.750 / 0.874 / 0.910", "—", "三种子一致"],
        ["未去泄漏对照臂（同环境复训）", 0.050, 0.931, 0.490, "仅对照"],
        ["　逐种子 G1", "0.072 / −0.069 / 0.000", "—", "—", "见过答案反而崩塌"],
        ["链感知消融臂（否决）", 0.278, 0.786, 0.532, "设计文档 §2.6"],
        ["历史对照：去泄漏（首轮旧实例）", 0.585, 0.842, 0.714, "Group 1 跨环境波动大"],
        ["历史对照：未去泄漏（首轮旧实例）", 0.457, 0.953, 0.705, "—"],
    ],
    note="指标：预测分数与公开 −log Kd 的组内 Spearman（公开真值仅用于独立自评，不进入产出链路）。",
    widths=[38, 24, 20, 10, 26],
)

wb.save(OUT)
print("wrote", OUT)
